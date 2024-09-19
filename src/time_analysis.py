import openpyxl
import os
import re
import statistics

output_excel_path = f'../../../analysis/time_output.xlsx'

# Check if the Excel file exists
if os.path.isfile(output_excel_path):
    wb = openpyxl.load_workbook(output_excel_path)
else:
    wb = openpyxl.Workbook()

sheet = wb.active

# Check if the header row is already present; if not, add it
if sheet['A1'].value is None:
    sheet.append([
        "Analysis Counter", "Node", "Read time", "Write time", 
        "Embedding time", "Choose action", "Reward calculate", 
        "Execution time", "Suppression time", "Training time"
    ])

print("Script running!!!")

def extract_values(line):
    pattern = re.compile(
        r'Analysis Counter:\s*(\d+)\s*'
        r'Node:\s*(\w+)\s*'
        r'Read time:\s*([\d\.]+)\s*'
        r'Write time:\s*([\d\.]+)\s*'
        r'Embedding time:\s*([\d\.]+)\s*'
        r'Choose action:\s*([\d\.]+)\s*'
        r'Reward calculate:\s*([\d\.]+)\s*'
        r'Execution time:\s*([\d\.]+)\s*'
        r'Suppression time:\s*([\d\.]+)\s*'
        r'Training time:\s*([\d\.]+)'
    )
    match = pattern.search(line)
    if match:
        return match.groups()
    return None

def calculate_averages(sheet):
    averages = ["Average", "", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    for col in range(3, 13):  # Columns C to M (3 to 13)
        values = []
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=col).value
            try:
                if cell_value is not None:
                    values.append(float(cell_value))
            except ValueError:
                continue
        if values:
            averages[col - 1] = statistics.mean(values)
    return averages

for sta_number in range(1, 9):
    log_file_path = f'/tmp/minindn/sta{sta_number}/reinforcement-sta{sta_number}.log'
    with open(log_file_path, 'r') as infile:
        for line in infile:
            if "Analysis Counter" in line:
                values = extract_values(line)
                if values:
                    sheet.append(values)
                    wb.save(output_excel_path)
    
    # Calculate averages after processing each log file
    averages = calculate_averages(sheet)
    sheet.append(averages)
    wb.save(output_excel_path)
    print(f"Averages after processing log file for sta{sta_number}: {averages}")
    cat_file_path = f'/tmp/minindn/sta{sta_number}/catchunks-sta1.txt-transfer9.dat.log'
    if os.path.isfile(cat_file_path): 
        print(sta_number)
        time_elapsed = ""
        goodput = ""
        timeout = ""
        retransmitted = ""
        rtt = ""
        rtt_min = ""
        rtt_avg = ""
        rtt_max = ""
        with open(cat_file_path, "r") as cat_file:
            for line in cat_file:
                if "Time elapsed: " in line:
                    time_elapsed = float(line.split(' ')[2])
                elif "Goodput: " in line:
                    goodput = float(line.split(' ')[1])
                elif "Timeouts:" in line:
                    timeout = float(line.split(' ')[1])
                elif "Retransmitted segments:" in line:
                    retransmitted = line.split(' 2')
                elif "RTT min/avg/max" in line:
                    rtt = line.split(' ')[3]
                    rtt_min = float(rtt.split('/')[0])
                    rtt_avg = float(rtt.split('/')[1])
                    rtt_max = float(rtt.split('/')[2])
            print("Time elapsed", sta_number, time_elapsed, goodput, timeout, rtt_min, rtt_avg, rtt_max)
            sheet.append(("Time elapsed", sta_number, time_elapsed, goodput, timeout, rtt_min, rtt_avg, rtt_max))
            wb.save(output_excel_path)

print("Script completed!")
