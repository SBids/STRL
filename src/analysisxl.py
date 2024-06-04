import openpyxl
import os
output_excel_path = f'../../../analysis/Xlfile/analysis1.xlsx'
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet.append(["Node", "Interest Rec", "Interest Sent","Data Rec", "Data Sent", "Time elapsed", "Goodput", "Timeout", "Retransmitted segment", "RTT"])
# Check if the Excel file exists
if os.path.isfile(output_excel_path):
    wb = openpyxl.load_workbook(output_excel_path)
else:
    wb = openpyxl.Workbook()

sheet = wb.active

# Check if the header row is already present; if not, add it
if sheet['A1'].value is None:
    sheet.append(["Node", "Interest Rec", "Interest Sent", "Data Rec", "Data Sent", "Time elapsed", "Goodput", "Timeout", "RTT min", "RTT avg", "Rtt max"])


print("Script running!!!")
for sta_number in range(1,7):
    data_tuples = []
    log_file_path = f'../../../analysis/rfiltered_lines{sta_number}.txt'
    
    cat_file_path = f'/tmp/minindn/sta{sta_number}/catchunks-sta1.txt.log'

    with open(log_file_path, "r") as log_file:
        data_rec_count = 0
        interest_sent_count = 0
        data_sent_finally = 0
        interest_rec_count = 0
        time_elapsed = ""
        goodput = ""
        timeout = ""
        retransmitted = ""
        rtt = ""
        rtt_min = ""
        rtt_avg = ""
        rtt_max = ""
        for line in log_file:
            if "Multicast data received" in line:
                data_rec_count += 1
            elif "Interest sent finally" in line:
                interest_sent_count += 1
            elif "Data sent finally" in line:
                data_sent_finally += 1
            elif "Multicast interest received" in line:
                interest_rec_count += 1 
    if os.path.isfile(cat_file_path): 
        with open(cat_file_path, "r") as cat_file:
            for line in cat_file:
                if "Time elapsed: " in line:
                    time_elapsed = float(line.split(' ')[2])
                elif "Goodput: " in line:
                    goodput = float(line.split(' ')[1])
                elif "Timeouts:" in line:
                    timeout = float(line.split(' ')[1])
                elif "Retransmitted segments:" in line:
                    retransmitted = line.split(' 2')
                elif "RTT min/avg/max" in line:
                    rtt = line.split(' ')[3]
                    rtt_min = float(rtt.split('/')[0])
                    rtt_avg = float(rtt.split('/')[1])
                    rtt_max = float(rtt.split('/')[2])
           

    data_tuples.append((sta_number, interest_rec_count, interest_sent_count, data_rec_count, data_sent_finally, time_elapsed, goodput, timeout, rtt_min, rtt_avg, rtt_max))
    print(data_tuples)
    for data_tuple in data_tuples:
        sheet.append(data_tuple)
    wb.save(output_excel_path)
    print("Log added for ", sta_number)
      
        

   
